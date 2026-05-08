from openpyxl import load_workbook
import pandas as pd
import os

path_xlsx = r"c:/Users/User/Documents/RappiReportes_version1/archivoExcel_PedidosYa_Español/orderDetails (4).xlsx"
path_xls = r"c:/Users/User/Documents/RappiReportes_version1/archivoExcel_PedidosYa_Español/invoice-9601391391.xls"

print("--- TESTING DIRECT OPENPYXL READ_ONLY ---")
try:
    wb = load_workbook(path_xlsx, read_only=True, data_only=True)
    ws = wb.active
    print("SUCCESS: Loaded workbook with read_only=True")
    print("First 5 rows:")
    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
        print(f"Row {i}: {row}")
except Exception as e:
    print(f"FAILED read_only: {e}")


print("\n--- INSPECTING XLS FILE HEADERS ---")
try:
    # Use pandas default (likely xlrd or similar)
    df = pd.read_excel(path_xls, header=None, nrows=15)
    print("XLS Rows:")
    for i, row in df.iterrows():
        # Clean print
        vals = [str(c).strip() for c in row if pd.notna(c) and str(c).strip()]
        if vals:
             print(f"Row {i}: {vals}")
except Exception as e:
    print(f"FAILED XLS: {e}")
